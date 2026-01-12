from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import List
from models import Detection
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


class ExcelReportGenerator:
    """Générateur de rapports Excel pour les détections"""
    
    @staticmethod
    def generate_report(detections: List[Detection], output_path: str):
        """Générer un rapport Excel avec les détections"""
        wb = Workbook()
        
        # Feuille 1: Vue par spot
        ExcelReportGenerator._create_sheet_by_spot(wb, detections)
        
        # Feuille 2: Vue par chaîne
        ExcelReportGenerator._create_sheet_by_chaine(wb, detections)
        
        # Feuille 3: Vue détaillée
        ExcelReportGenerator._create_sheet_detailed(wb, detections)
        
        wb.save(output_path)
        logger.info(f"Rapport Excel généré: {output_path}")
    
    @staticmethod
    def _create_sheet_by_spot(wb: Workbook, detections: List[Detection]):
        """Feuille: Résumé par spot"""
        ws = wb.active
        ws.title = "Par Spot"
        
        # En-têtes
        headers = ["Spot", "Nb Détections", "Nb Chaînes", "Confiance Moy.", "Première Diffusion", "Dernière Diffusion"]
        ws.append(headers)
        ExcelReportGenerator._style_header_row(ws, 1)
        
        # Grouper par spot
        by_spot = {}
        for d in detections:
            if d.spot_nom not in by_spot:
                by_spot[d.spot_nom] = []
            by_spot[d.spot_nom].append(d)
        
        # Données
        for spot_nom, spot_detections in sorted(by_spot.items()):
            chaines = set(d.enreg_chaine_nom for d in spot_detections)
            avg_conf = sum(d.confidence for d in spot_detections) / len(spot_detections)
            dates = sorted([d.enreg_date for d in spot_detections])
            
            ws.append([
                spot_nom,
                len(spot_detections),
                len(chaines),
                avg_conf,
                dates[0],
                dates[-1]
            ])
        
        ExcelReportGenerator._format_columns(ws, [30, 15, 12, 15, 18, 18])
        ExcelReportGenerator._apply_number_formats(ws, {
            'D': '0.0"%"',
            'E': 'yyyy-mm-dd',
            'F': 'yyyy-mm-dd'
        })
    
    @staticmethod
    def _create_sheet_by_chaine(wb: Workbook, detections: List[Detection]):
        """Feuille: Résumé par chaîne"""
        ws = wb.create_sheet("Par Chaîne")
        
        headers = ["Chaîne", "Nb Détections", "Nb Spots Uniques", "Confiance Moy."]
        ws.append(headers)
        ExcelReportGenerator._style_header_row(ws, 1)
        
        # Grouper par chaîne
        by_chaine = {}
        for d in detections:
            if d.enreg_chaine_nom not in by_chaine:
                by_chaine[d.enreg_chaine_nom] = []
            by_chaine[d.enreg_chaine_nom].append(d)
        
        for chaine, chaine_detections in sorted(by_chaine.items()):
            spots = set(d.spot_nom for d in chaine_detections)
            avg_conf = sum(d.confidence for d in chaine_detections) / len(chaine_detections)
            
            ws.append([
                chaine,
                len(chaine_detections),
                len(spots),
                avg_conf
            ])
        
        ExcelReportGenerator._format_columns(ws, [15, 15, 18, 15])
        ExcelReportGenerator._apply_number_formats(ws, {'D': '0.0"%"'})
    
    @staticmethod
    def _create_sheet_detailed(wb: Workbook, detections: List[Detection]):
        """Feuille: Détails complets"""
        ws = wb.create_sheet("Détails")
        
        headers = [
            "Spot", "Chaîne", "Date", "Enregistrement",
            "Début", "Fin", "Durée (s)", "Type", "Confiance"
        ]
        ws.append(headers)
        ExcelReportGenerator._style_header_row(ws, 1)
        
        for d in sorted(detections, key=lambda x: (x.enreg_date, x.enreg_chaine_nom, x.start_seconds)):
            duree = d.end_seconds - d.start_seconds
            ws.append([
                d.spot_nom,
                d.enreg_chaine_nom,
                d.enreg_date,
                d.enreg_nom,
                d.start_time,
                d.end_time,
                duree,
                "EXACT" if d.match_type == "exact" else "FUZZY",
                d.confidence
            ])
        
        ExcelReportGenerator._format_columns(ws, [25, 10, 12, 35, 12, 12, 12, 10, 12])
        ExcelReportGenerator._apply_number_formats(ws, {
            'C': 'yyyy-mm-dd',
            'G': '0.0',
            'I': '0.0"%"'
        })
    
    @staticmethod
    def _style_header_row(ws, row_num: int):
        """Styliser la ligne d'en-tête"""
        for cell in ws[row_num]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    @staticmethod
    def _format_columns(ws, widths: List[int]):
        """Définir les largeurs de colonnes"""
        for idx, width in enumerate(widths, start=1):
            col_letter = ws.cell(1, idx).column_letter
            ws.column_dimensions[col_letter].width = width
    
    @staticmethod
    def _apply_number_formats(ws, formats: dict):
        """Appliquer des formats numériques par colonne"""
        for col_letter, fmt in formats.items():
            for row in range(2, ws.max_row + 1):
                ws[f"{col_letter}{row}"].number_format = fmt
